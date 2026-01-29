#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract competency model data from Excel file.

Extracts:
- Employee categories (5)
- Competency blocks and clusters
- Competencies
- Target levels per (category × competency)
- Textual descriptions of levels 1–5
- Glossary
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook

def normalize_key(text):
    """Normalize text to create canonical keys for matching."""
    if not text:
        return ""
    # Remove extra whitespace, convert to lowercase, replace spaces with underscores
    # Remove special characters that might cause issues
    text = str(text).strip()
    # Replace common Russian characters and normalize
    text = text.lower()
    # Replace spaces and special chars with underscores
    import re
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '_', text)
    return text.strip('_')

def build_merged_value_map(ws):
    """Build map of merged cell coordinates -> top-left value."""
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_map[(row, col)] = top_left
    return merged_map

def get_cell_value(ws, row, col, merged_map):
    """Get cell value, filling from merged cells when needed."""
    value = ws.cell(row=row, column=col).value
    if value is None and merged_map:
        value = merged_map.get((row, col))
    return value

def find_categories_row(ws, merged_map):
    """Try to detect the header row with category names in columns G-K."""
    best_row = None
    best_score = -1
    best_values = []

    max_scan = min(10, ws.max_row)
    for row_idx in range(1, max_scan + 1):
        values = [get_cell_value(ws, row_idx, col, merged_map) for col in range(7, 12)]
        score = 0
        cleaned = []
        non_empty = []
        for value in values:
            if value is None:
                cleaned.append("")
                continue
            text = str(value).strip()
            if not text or text == "None":
                cleaned.append("")
                continue
            try:
                float(text)
                cleaned.append("")
                continue
            except ValueError:
                cleaned.append(text)
                score += 1
                non_empty.append(text)

        if score <= 0:
            continue

        row_text = " ".join(
            [str(get_cell_value(ws, row_idx, col, merged_map) or "") for col in range(1, 7)]
        ).lower()
        if "категор" in row_text:
            score += 1
        unique_count = len({val for val in non_empty if val})
        score += unique_count
        if unique_count == 1 and score > 0:
            score -= 2

        if score > best_score:
            best_score = score
            best_row = row_idx
            best_values = cleaned

    if best_row is None:
        return 2, []
    return best_row, best_values

def extract_main_matrix(ws):
    """Extract competency data from main matrix sheet."""
    categories = []
    competencies = []
    blocks = {}
    clusters = {}
    target_levels = {}
    level_descriptions = {}
    
    merged_map = build_merged_value_map(ws)

    # Read header row to identify categories
    header_row, header_values = find_categories_row(ws, merged_map)
    if header_values:
        categories = [cat for cat in header_values if cat and cat != "None"]
    else:
        # Fallback to row 2, columns G-K
        categories = []
        for col in range(7, 12):
            cell_value = get_cell_value(ws, 2, col, merged_map)
            if cell_value:
                categories.append(str(cell_value).strip())
        categories = [cat for cat in categories if cat and cat != "None"]
    
    print(f"Found categories: {categories}")
    
    # Track current block and cluster (they're merged cells, so we need to track them)
    current_block = None
    current_cluster = None
    
    # Process data rows (starting from row after header)
    start_row = (header_row or 2) + 1
    for row_idx in range(start_row, ws.max_row + 1):
        row = [get_cell_value(ws, row_idx, col, merged_map) for col in range(1, ws.max_column + 1)]
        # Check if row is empty
        if not any(row[:6]):  # First 6 columns should have data
            continue
        
        # Extract block (column A)
        block_name = str(row[0]).strip() if row[0] else None
        if block_name and block_name != "None":
            current_block = block_name
            block_id = normalize_key(current_block)
            if block_id not in blocks:
                blocks[block_id] = {
                    "id": block_id,
                    "name": current_block,
                    "clusters": []
                }
        
        # Extract cluster (column B)
        cluster_name = str(row[1]).strip() if row[1] else None
        if cluster_name and cluster_name != "None":
            current_cluster = cluster_name
            cluster_id = normalize_key(current_cluster)
            if cluster_id not in clusters:
                clusters[cluster_id] = {
                    "id": cluster_id,
                    "name": current_cluster,
                    "block_id": normalize_key(current_block) if current_block else "",
                    "competencies": []
                }
                if current_block:
                    block_id = normalize_key(current_block)
                    if block_id in blocks:
                        if cluster_id not in blocks[block_id]["clusters"]:
                            blocks[block_id]["clusters"].append(cluster_id)
        
        # Extract competency (column C)
        competency_name = str(row[2]).strip() if row[2] else None
        if not competency_name or competency_name == "None":
            continue
        
        competency_id = normalize_key(competency_name)
        
        # Extract other fields
        description = str(row[3]).strip() if row[3] else ""
        required_skills = str(row[4]).strip() if row[4] else ""
        priority = str(row[5]).strip() if row[5] else ""
        
        # Extract target levels for each category (columns G-K, indices 6-10)
        category_levels = {}
        for cat_idx, category in enumerate(categories):
            if cat_idx < len(row[6:11]):
                level_value = row[6 + cat_idx]
                if level_value and str(level_value).strip() and str(level_value) != "None":
                    try:
                        level = int(float(str(level_value).strip()))
                        category_levels[normalize_key(category)] = level
                    except (ValueError, TypeError):
                        pass
        
        # Extract level descriptions (columns L-P, indices 11-15)
        level_descs = {}
        base_desc_idx = 11  # Column L (index 11)
        for level_num in range(1, 6):
            desc_idx = base_desc_idx + (level_num - 1)
            if desc_idx < len(row):
                desc = str(row[desc_idx]).strip() if row[desc_idx] else ""
                if desc and desc != "None":
                    level_descs[str(level_num)] = desc
        
        # Create competency object
        competency = {
            "id": competency_id,
            "name": competency_name,
            "description": description,
            "required_skills": required_skills,
            "priority": priority,
            "block_id": normalize_key(current_block) if current_block else "",
            "cluster_id": normalize_key(current_cluster) if current_cluster else "",
            "target_levels": category_levels,
            "level_descriptions": level_descs
        }
        
        # Ensure block exists before referencing
        if current_block:
            block_id = normalize_key(current_block)
            if block_id not in blocks:
                blocks[block_id] = {
                    "id": block_id,
                    "name": current_block,
                    "clusters": []
                }
        
        competencies.append(competency)
        
        # Add to cluster
        if current_cluster:
            cluster_key = normalize_key(current_cluster)
            if cluster_key in clusters:
                clusters[cluster_key]["competencies"].append(competency_id)
        
        # Store target levels by category
        for cat_key, level in category_levels.items():
            if cat_key not in target_levels:
                target_levels[cat_key] = {}
            target_levels[cat_key][competency_id] = level
        
        # Store level descriptions
        if level_descs:
            level_descriptions[competency_id] = level_descs
    
    # Convert blocks and clusters to lists
    blocks_list = list(blocks.values())
    clusters_list = list(clusters.values())
    
    return {
        "categories": [{"id": normalize_key(cat), "name": cat} for cat in categories],
        "blocks": blocks_list,
        "clusters": clusters_list,
        "competencies": competencies,
        "target_levels": target_levels,
        "level_descriptions": level_descriptions
    }

def extract_glossary(ws):
    """Extract glossary from glossary sheet."""
    glossary = {}
    
    # Skip header row, process data
    for row in ws.iter_rows(min_row=2, values_only=True):
        term = str(row[0]).strip() if row[0] else None
        definition = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        
        if term and term != "None":
            term_key = normalize_key(term)
            glossary[term_key] = {
                "term": term,
                "definition": definition
            }
    
    return glossary

def extract_level_scale(ws):
    """Extract level scale descriptions from 'Шкала развития компетенций' sheet."""
    scale = {}
    
    # This sheet likely has general level descriptions
    # Structure may vary, so we'll try to extract it
    for row in ws.iter_rows(min_row=2, values_only=True):
        level = str(row[0]).strip() if row[0] else None
        description = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        
        if level and level != "None":
            scale[level] = description
    
    return scale

def main():
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / "data-src" / "model.xlsx"
    if not excel_path.exists():
        excel_path = base_dir / "data" / "Модель_цифровых_компетенций.xlsx"
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    print(f"Found sheets: {wb.sheetnames}")
    
    all_data = {}
    
    # Extract from main matrix sheet
    main_sheet_name = "Матрица цифровых компетенций"
    if main_sheet_name in wb.sheetnames:
        print(f"\nExtracting from main sheet: {main_sheet_name}")
        main_data = extract_main_matrix(wb[main_sheet_name])
        all_data.update(main_data)
    else:
        print(f"Warning: Main sheet '{main_sheet_name}' not found")
        # Try first sheet
        if wb.sheetnames:
            print(f"Trying first sheet: {wb.sheetnames[0]}")
            main_data = extract_main_matrix(wb[wb.sheetnames[0]])
            all_data.update(main_data)
    
    # Extract glossary
    glossary_sheet_name = "Словарь терминов и сокращений"
    if glossary_sheet_name in wb.sheetnames:
        print(f"\nExtracting glossary from: {glossary_sheet_name}")
        glossary = extract_glossary(wb[glossary_sheet_name])
        all_data["glossary"] = glossary
        print(f"Extracted {len(glossary)} glossary terms")
    
    # Extract level scale
    scale_sheet_name = "Шкала развития компетенций"
    if scale_sheet_name in wb.sheetnames:
        print(f"\nExtracting level scale from: {scale_sheet_name}")
        scale = extract_level_scale(wb[scale_sheet_name])
        all_data["level_scale"] = scale
    
    # Save output
    output_path = Path(__file__).parent.parent / "frontend" / "data" / "model.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ Extracted data saved to: {output_path}")
    print(f"  - Categories: {len(all_data.get('categories', []))}")
    print(f"  - Blocks: {len(all_data.get('blocks', []))}")
    print(f"  - Clusters: {len(all_data.get('clusters', []))}")
    print(f"  - Competencies: {len(all_data.get('competencies', []))}")
    print(f"  - Glossary terms: {len(all_data.get('glossary', {}))}")
    
    return all_data

if __name__ == "__main__":
    main()
