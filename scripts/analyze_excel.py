#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyze Excel structure to understand data layout.
"""

from pathlib import Path
from openpyxl import load_workbook

def analyze_sheet(ws, max_rows=50, max_cols=20):
    """Print structure of a worksheet."""
    print(f"\n{'='*80}")
    print(f"Sheet: {ws.title}")
    print(f"{'='*80}")
    
    # Print header rows
    print("\nFirst 20 rows (first 15 columns):")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, max_rows), 
                                         min_col=1, max_col=min(15, max_cols), 
                                         values_only=True), 1):
        row_str = " | ".join([str(cell)[:30] if cell is not None else "" for cell in row])
        print(f"Row {i:2d}: {row_str}")
    
    # Check for merged cells
    if ws.merged_cells.ranges:
        print(f"\nMerged cells: {len(ws.merged_cells.ranges)}")
        for merged in list(ws.merged_cells.ranges)[:10]:
            print(f"  {merged}")

def main():
    excel_path = Path(__file__).parent.parent / "data-src" / "model.xlsx"
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    print(f"\nTotal sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        analyze_sheet(ws)

if __name__ == "__main__":
    main()
